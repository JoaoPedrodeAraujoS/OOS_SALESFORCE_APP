import pandas as pd
import numpy as np
import xlwings as xw
import os, os.path, sys
from datetime import datetime
from dateutil.relativedelta import relativedelta
import re
from dateutil.parser._parser import ParserError
from api import read_preferences
from sqlalchemy import create_engine, null
from openpyxl import load_workbook
from datetime import date, timedelta
import xml.etree.ElementTree as et
from win32com.client import Dispatch
import warnings

class Parser:

    def __init__(self, path, file_pattern):
        self.path = path
        self.file_pattern = file_pattern

    def get_list_files(self):
        cleaned_path = []
        preferences = read_preferences()
        from_date = datetime.strptime(preferences['date_entry_upload_since'], '%d/%m/%Y')
        # discovering just the necessary folders/files
        # Reduced 8x the time in comparison to simple os.walk in root path
        list_dates = self.get_dates_until(from_date, datetime.now()+relativedelta(months=1))
        walkers = [os.walk(
            self.path + '/' + str(data.year) + '/' + str(data.month).zfill(2)
        ) for data in list_dates]
        
        for walker in walkers:
            for root, dirname, fnames in walker:
                for fname in fnames:
                    if fname.upper().startswith(self.file_pattern):
                        cleaned_path.append(
                            (root + '/' + fname).replace('\\', '/'))

        return cleaned_path
    
    @staticmethod
    def get_dates_until(from_date, to_date):
        total_dates = []
        while from_date <= to_date:
            total_dates.append(from_date.date())
            from_date += relativedelta(months=1)
        return total_dates
    
    @staticmethod
    def discover_parser(fname):
        if 'AZUL' in fname:
            return AzulParser
        elif 'WIDEROE'  in fname:
            return WideroeParser
        elif 'AIR ASTANA' in fname:
            return AstanaParser
        elif 'HELVETIC AIRWAYS' in fname:
            return HelveticParser
        elif 'BELAVIA' in fname:
            return BelaviaParser        
        elif 'KLM CITYHOPPER' in fname:
            return KLMParser
        else:
            raise (TypeError, 'Did not found any implemented parser to this filename')
            
    
    @staticmethod
    def _skip_blank_header(df):
        df = df.dropna(how='all').reset_index(drop=True)
        while sum([not (pd.isna(col) or 'Unnamed' in col) for col in df.columns]) < 5:
            df.columns = df.iloc[0]
            df = df.drop([0], axis=0).reset_index(drop=True)
        
        return df
    
    @staticmethod
    def _commom_clean(df):
        df['Action_Description__c'] = df['Action_Description__c'].str[:40000].str.replace(
                '\n', '<br />').replace('\r', '')
        df['Event_Description__c'] = df['Event_Description__c'].str[:40000].str.replace(
                '\n', '<br />').replace('\r', '')
        
        df[['Start_Date__c', 'Start_Time__c']] = df['Start_Date__c'].str[:-1].str.split(
            'T', 1, expand=True).fillna('')
        
        df[['Release_Date__c', 'Release_Time__c']] = df['Release_Date__c'].str[:-1].str.split(
            'T', 1, expand=True).fillna('')
        
        return df

    @staticmethod
    def _date_format_f(cell):
        if isinstance(cell, datetime):
            return datetime.strftime(cell.date(), '%Y-%m-%d')
        return cell

    @staticmethod
    def normalize_datetime(df, c_date, c_time=None):
        df[c_date] = df[c_date].fillna('01/01/2000').apply(
            Parser._date_format_f)
        if c_time is not None:
            df[c_time] = df[c_time].fillna('00:00:00').astype(
                str).str.extract(r'([0-9]{1,2}\:[0-9]{1,2}(?:\:[0-9]{1,2})?)')
            s = df[c_date] + ' ' + df[c_time]
        else:
            s = df[c_date]

        dummy_date = '01/01/2000 00:00:00'
        s = pd.to_datetime(s, dayfirst=True).apply(
            datetime.isoformat) + 'Z'
        s = s.replace(datetime.isoformat(
            pd.to_datetime(dummy_date)) + 'Z', '')

        return s
    
    
    @staticmethod
    def _normalize_oos_time_by_row(s, c_oos_time, c_start_date, c_end_date):
        if pd.isna(s[c_oos_time]):
            start_date = pd.to_datetime(s[c_start_date][:-1])
            end_date = pd.to_datetime(s[c_end_date][:-1])
            
            return (end_date - start_date).total_seconds() / 3600
        else:
            try:
                return float(s[c_oos_time])
            except ValueError:
                pass
            
            try:
                date = pd.to_datetime(s[c_oos_time])
                if date.date() != datetime.now().date():
                   return ((date.month - 1) * 31 * 24) + (date.day * 24) + date.hour + (date.minute / 60)
            except TypeError:
                pass
            except ParserError:
                pass
            
            try:
                cell_splitted = str(s[c_oos_time]).split(':')
                if len(cell_splitted) > 1:
                    return float(cell_splitted[0]) + float(cell_splitted[1]) / 60
            except TypeError:
                pass
            
            return s[c_oos_time]
        
    
    def normalize_oos_time(self, df, cleaned_df, c_oos_time):
        if c_oos_time is None:
            cleaned_df['OOS_Total_Time__c'] = np.nan
        else:
            cleaned_df['OOS_Total_Time__c'] = df[c_oos_time].apply(
                lambda x: pd.NA if pd.isna(x) else str(x).strip()
            )  

        return cleaned_df.apply(
            self._normalize_oos_time_by_row,
            args=('OOS_Total_Time__c', 'Start_Date__c', 'Release_Date__c'),
            axis=1
        )

    def get_unprocessed_files(self):
        need_analysis = []
        paths = self.get_list_files()
        #print(paths)
        '''
        if os.path.isfile('settings/history_files.txt'):
            with open('settings/history_files.txt', 'r') as f:
                history_paths = f.read().split('\n')

            for path in paths:
                if path.replace('ø', 'o').replace('@', 'a') not in history_paths:
                    need_analysis.append(path)
            #print(need_analysis)
        else:
        '''
        for path in paths:    
            need_analysis.append(path)
        
        

        return need_analysis
    
    def load_file(self, path, converters=None, dtype=None):
        df = pd.read_excel(path, converters=converters, dtype=dtype)
        df = self._skip_blank_header(df)
        df.columns = [col.strip().lower() for col in df.columns]
        return df
    
    def load_file_azul(self, path, converters=None, dtype=None, sheet_name=None):
        df = pd.read_excel(path, converters=converters, dtype=dtype, sheet_name='OOS')
        df = self._skip_blank_header(df)
        df.columns = [col.strip().lower() for col in df.columns]
        return df

    def load_file_astana(self, path, converters=None, dtype=None):
        df = pd.read_excel(path, converters=converters, dtype=dtype, skiprows=2)
        df = self._skip_blank_header(df)
        df.columns = [col.strip().lower() for col in df.columns]
        return df
    
    def load_file_belavia(self, path, converters=None, dtype=None, sheet_name=None):
        df = pd.read_excel(path, converters=converters, dtype=dtype, sheet_name='OUT-OF-SERVICE')
        df = self._skip_blank_header(df)
        df.columns = [col.strip().lower() for col in df.columns]
        return df

    def load_file_klm(self, path, converters=None, dtype=None, sheet_name=None):
        df = pd.read_excel(path, converters=converters, dtype=dtype, sheet_name='OOS')
        df = self._skip_blank_header(df)
        df.columns = [col.strip().lower() for col in df.columns]
        return df
    def load_file_klm_pirep(self, path, converters=None, dtype=None, sheet_name=None):
        df = pd.read_excel(path, converters=converters, dtype=dtype, sheet_name='Pirep')
        df = self._skip_blank_header(df)
        df.columns = [col.strip().lower() for col in df.columns]
        return df
    def get_cleaned_df(self, fname):
        raise NotImplementedError

    @staticmethod
    def get_reference_date(fname, length):
        month = None
        year = None
        for part in fname.split('/'):
            if part.isdigit():
                if int(part) > 0 and int(part) <= 12:
                    month = part
                elif int(part) > 2000 and int(part) < 2200:
                    year = part

        return length * ['{}-{}-{}'.format(year, month, '01')]

 #FUNCTION FOR EVENT CORRELATE INTER AND OOS FOR KLM CITYHOOPPER    
    @staticmethod
    def merge_klm(self):
        from datetime import date, timedelta
        warnings.filterwarnings('ignore')
        print('Executando o relacionamento para a KLM CITYHOPPER...')
        
        def get_fname(preferences):
            ano = preferences["date_entry_upload_since"].split('/')[2]
            mes = preferences["date_entry_upload_since"].split('/')[1]
            fname = '//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/1 - Dados recebidos/4 - EMEA/KLM CITYHOPPER' + '/' + ano + '/' + mes + '/' + 'OOS_DATA-KLM.xlsx'
            return fname

        def _skip_blank_header(df):
            df = df.dropna(how='all').reset_index(drop=True)
            while sum([not (pd.isna(col) or 'Unnamed' in col) for col in df.columns]) < 5:
                df.columns = df.iloc[0]
                df = df.drop([0], axis=0).reset_index(drop=True)
        
            return df        
        
        def load_file_klm_pirep(self, preferences, fname, converters=None, dtype=None, sheet_name=None):
            fname = get_fname(preferences)            
            df = pd.read_excel(fname, converters=converters, dtype=dtype, sheet_name=5)
            df = _skip_blank_header(df)
            df.columns = [col.strip().lower() for col in df.columns]
            return df
        
        #DEFINE THE DIRECTORY OF THE FILE
        preferences = read_preferences()
        fname = get_fname(preferences)
        
        
        def Return_SQL():
            return ("""
                SELECT
                    COD_SEQ_PROB AS "Problem",   
                    REG_AENV_DATA_EVNT AS "Register",
                    NOM_SIMPL_ORGZ_DATA_EVNT AS "Operator (Simplified Name)",
                    COD_CAP_ATA_FALHA  AS "Ata Chapter",   
                    COD_SUCP_ATA_FALHA AS "Ata Sub. Chapter",  
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 
                    'PROB:', 1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 
                    'ACTION:', 1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 1, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) -1)  ELSE  DSC_ACAO_CRRT  END AS "Event Description", 
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 'PROB:', 
                    1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1))  ELSE  DSC_ACAO_CRRT  END AS "Action Description",
                    to_char(DAT_EVNT,'MM/dd/yyyy') AS "Event Date"
                FROM
                    RCP_VW_PROB_AERONAVE wp 
                    WHERE 1 = 1   AND DAT_REF >= TO_DATE(:DATE_REF,'DD/MM/YYYY')  
                    AND DAT_REF <= TO_DATE(:DATE_REF,'DD/MM/YYYY')
                    AND COD_TIPO_PROB = 'INTER'
                    AND COD_MERC_AVC = 'C'   AND COD_PROJ IN ('1902','1952')   
                    AND COD_ORGZ_DATA_EVNT = 9971

            """)

        #FUNNCTION FOR FORMATING THE PIREP TABLE OF KLM
        def parser_sheet_pirep_klm(fname):
            wb = load_workbook(filename = fname)
            sheets = wb.sheetnames
            #NUMBER OF TABLE
            n = 5
            ws = wb[sheets[n]]
            ws.merge_cells('A7:A8')
            ws['A7'].value = 'Report Date'
            ws.merge_cells('D7:D8')
            ws['D7'].value = 'Aircraft Registration'
            ws.merge_cells('E7:E8')
            ws['E7'].value = 'Aircraft type'
            ws.merge_cells('F7:F8')
            ws['F7'].value = 'Tech Log Entry'
            wb.save(filename = fname)
                    
        def get_cleaned_df(self, fname):
            parser_sheet_pirep_klm(fname)
            preferences = read_preferences()
            fname = get_fname(preferences)
            
            converters = {
                'A/C': (lambda x: 'PH-' + x.strip()),
                'Remarks': (lambda x: x.upper())}
            
            df = pd.read_excel(fname, converters=converters, sheet_name='OOS')
            df_filter = pd.DataFrame(df)
            df_filter = df[df['AC Type'] == 'E195E2']
            df_filter['ATA'] = df_filter['ATA'].map(str)
            df_filter['ATA'] = df_filter['ATA'].str.lstrip('0')
            df_filter['ATA'] = df_filter['ATA'].str.strip()
            df_filter_proc_pirep =  df_filter.filter(items = ['A/C', 'ATA', 'Remarks'])
            df_filter_proc_pirep['report date'] = pd.to_datetime(df_filter['Start Date'], format='%d-%m-%Y')
            df_filter_proc_pirep['report date'] = df_filter_proc_pirep['report date'].map(str)
            df_filter_proc_pirep['report date'] = df_filter_proc_pirep['report date'].str.strip()
            ac_formated = df_filter_proc_pirep['A/C'].str.split("-", n=1, expand=True)
            df_filter_proc_pirep['A/C_L'] = ac_formated[0]
            df_filter_proc_pirep['A/C_R'] = ac_formated[1]
            df_filter_proc_pirep['A/C'] = df_filter_proc_pirep['A/C_L'] + df_filter_proc_pirep['A/C_R']
            df_filter_proc_pirep['A/C'] = df_filter_proc_pirep['A/C'].str.strip()
            df_filter_proc_pirep['ATA'] = df_filter_proc_pirep['ATA'].str.strip()
            df_filter_proc_pirep = df_filter_proc_pirep.filter(items = ['A/C', 'ATA', 'Remarks', 'report date'])
            
            df_filter_proc =  df_filter.filter(items = ['A/C', 'ATA', 'Remarks'])
            df_filter_proc['report date'] = pd.to_datetime(df_filter['Start Date'], format='%d-%m-%Y')
            df_filter_proc['report date'] = df_filter_proc['report date'].map(str)
            df_filter_proc['report date'] = df_filter_proc['report date'].str.strip()
            df_filter_proc['A/C'] = df_filter_proc['A/C'].str.strip()
            df_filter_proc['ATA'] = df_filter_proc['ATA'].str.strip()
            
            df_pirep_filter = load_file_klm_pirep(self, preferences, fname)
            df_pirep_filter['report type'] = df_pirep_filter[df_pirep_filter['report type'] == 'Pilot']
            df_pirep_filter = df_pirep_filter.filter(items=['report date','ata', 'aircraft registration', 'cause text', 'fix text'])
            df_pirep_filter['aircraft registration'] = df_pirep_filter['aircraft registration'].map(str)
            df_pirep_filter['aircraft registration'] = df_pirep_filter['aircraft registration'].str.strip() 
            df_pirep_filter['ata'] = df_pirep_filter['ata'].map(str)       
            df_pirep_filter['ata'] = df_pirep_filter['ata'].str.strip()
            df_pirep_filter['report date'] = pd.to_datetime(df_pirep_filter['report date'].map(str), format='%Y-%m-%d')
            df_pirep_filter['report date'] = df_pirep_filter['report date'].map(str)
            df_pirep_filter['report date'] = df_pirep_filter['report date'].str.strip()
            
            #RCP CONFIGURATIONS
            conc = 'oracle://rcp_con:CDF34@ora-pr06flm:1527/pr06'
            engine = create_engine(conc)

            #GETTING THE YAER AND MONTH REFERENCE
            date_ref = preferences['date_entry_upload_since']
            date_ref = pd.to_datetime(date_ref, format='%d/%m/%Y')
            SQL = Return_SQL()
            print('RUNNING THE QUERY')
            dfquery = pd.read_sql_query(SQL, engine, params={'DATE_REF': date_ref.strftime("%d/%m/%Y")})
            print('QUERY EXECUTED')
            df_klm = pd.DataFrame(dfquery)
            df_klm['Event Date'] = pd.to_datetime(df_klm['Event Date'].map(str), format='%m/%d/%Y')
            df_klm['Event Date'] = df_klm['Event Date'].map(str)
            df_klm['Event Date'] = df_klm['Event Date'].str.strip()
            df_klm['Ata Chapter'] = df_klm['Ata Chapter'].map(str)
            df_klm['Ata Chapter'] = df_klm['Ata Chapter'].str.strip()

            #MERGE INTER KLM
            df_filter_inter = pd.merge(df_filter_proc, df_klm, left_on=['report date','A/C', 'ATA'], 
                                       right_on=['Event Date','Register', 'Ata Chapter'], 
                                       how='inner')

            #FILTRING THE COLUMNS NECESSARY (INTER)
            df_filter_inter = df_filter_inter.filter(items = ['Problem', 'Event Date', 'Register', 'Ata Chapter',
                                                              'Remarks', 'Event Description', 
                                                              'Action Description'])

            df_filter_inter = df_filter_inter.rename({'Event Date':'DATE_OOS',
                                                    'Register':'REGISTER_OOS',
                                                    'Event Description':'EVENT_INTER',
                                                    'Action Description':'ACTION_INTER',
                                                    'Remarks':'EVENT_OOS',
                                                    'Ata Chapter':'ATA_OOS'}, axis=1)
            #MERGE PIREP KLM
            df_filter_pirep = pd.merge(df_filter_proc_pirep, df_pirep_filter, left_on=['report date', 'ATA', 'A/C'], 
                                           right_on=['report date', 'ata', 'aircraft registration'], 
                                           how='inner')
            
            #FILTRING THE COLUMNS NECESSARY (PIREP)
            df_filter_pirep = df_filter_pirep.filter(items = ['report date', 'aircraft registration', 'ata', 
                                                              'Remarks', 'cause text', 'fix text'])

            df_filter_pirep = df_filter_pirep.rename({'report date':'DATE_OOS',
                                                    'aircraft registration':'REGISTER_OOS',
                                                    'cause text':'EVENT_PIREP',
                                                    'fix text':'ACTION_PIREP',
                                                    'Remarks':'EVENT_OOS',
                                                    'ata':'ATA_OOS'}, axis=1)

            #SAVING THE FILE IN THE DIRECTORY
            with pd.ExcelWriter("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/KLM/INTER_PIREP_KLM.xlsx", engine="xlsxwriter") as writer:
                df_filter_inter.to_excel(writer, sheet_name="INTER", index=False)
                df_filter_pirep.to_excel(writer, sheet_name="PIREP", index=False)
            writer.save()
            
            #FORMATING THE WORKSHEET
            print('Fromatando as abas')
            excel = Dispatch('Excel.Application')
            wb_formated = excel.Workbooks.Open("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/KLM/INTER_PIREP_KLM.xlsx")
            excel.Application.ScreenUpdating = False
            excel.Worksheets(1).Activate()
            excel.ActiveSheet.Columns("A:D").AutoFit()
            excel.ActiveSheet.Columns("B").NumberFormat = 'DD/MM/AAAA'
            excel.ActiveSheet.Columns("E").ColumnWidth = 19
            excel.ActiveSheet.Columns("F").ColumnWidth = 19
            excel.ActiveSheet.Columns("G").ColumnWidth = 19
            excel.Worksheets(2).Activate()
            excel.ActiveSheet.Columns("A:C").AutoFit()
            excel.ActiveSheet.Columns("A").NumberFormat = 'DD/MM/AAAA'
            excel.ActiveSheet.Columns("D").ColumnWidth = 19
            excel.ActiveSheet.Columns("E").ColumnWidth = 19
            excel.ActiveSheet.Columns("F").ColumnWidth = 19
            excel.Application.ScreenUpdating = True
            wb_formated.Save()
            wb_formated.Close()
            print('Formatação concluída!')

            print('Arbindo as planilhas')
            wb = xw.Book("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/KLM/INTER_PIREP_KLM.xlsx")
            print('Finalizado!')
        get_cleaned_df(self, fname)

#FUNCTION FOR EVENT CORRELATE INTER AND OOS FOR WIDEROE
    @staticmethod    
    def merge_wideroe(self):
        print('Executando o relacionamento para a WIDEROE...')
        from datetime import date, timedelta
        warnings.filterwarnings('ignore')
        def get_fname(preferences):

            #GETTING THE YAER AND MONTH REFERENCE
            ano = preferences["date_entry_upload_since"].split('/')[2]
            mes = preferences["date_entry_upload_since"].split('/')[1]
            fname = '//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/1 - Dados recebidos/4 - EMEA/WIDEROE' + '/' + ano + '/' + mes + '/' + 'OOS_DATA-WIDEROE.xlsx'
            return fname

        #DEFINE THE DIRECTORY OF THE FILE
        preferences = read_preferences()
        fname = get_fname(preferences)

        #SQL QUERY INTER WIDEROE
        def Return_SQL():
            
            return ("""
                SELECT
                    COD_SEQ_PROB AS "Problem",   
                    REG_AENV_DATA_EVNT AS "Register",
                    NOM_SIMPL_ORGZ_DATA_EVNT AS "Operator (Simplified Name)",
                    COD_CAP_ATA_FALHA  AS "Ata Chapter",   
                    COD_SUCP_ATA_FALHA AS "Ata Sub. Chapter",
                    NUM_FICSA AS "Fisca", 
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 
                    'PROB:', 1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 
                    'ACTION:', 1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 1, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) -1)  ELSE  DSC_ACAO_CRRT  END AS "Event Description", 
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 'PROB:', 
                    1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1))  ELSE  DSC_ACAO_CRRT  END AS "Action Description",
                    to_char(DAT_EVNT,'MM/dd/yyyy') AS "Event Date"
                FROM
                    RCP_VW_PROB_AERONAVE wp 
                    WHERE 1 = 1   AND DAT_REF >= TO_DATE(:DATE_REF,'DD/MM/YYYY')  
                    AND DAT_REF <= TO_DATE(:DATE_REF,'DD/MM/YYYY')
                    AND COD_TIPO_PROB = 'INTER'
                    AND COD_MERC_AVC = 'C'   AND COD_PROJ IN ('1902','1952')   
                    AND COD_ORGZ_DATA_EVNT = 10920
            """)
            
        def get_cleaned_df(self, fname):
            warnings.filterwarnings('ignore')
            preferences = read_preferences()
            fname = get_fname(preferences)
            
            #CONVERTING THE REGISTER TO THE FULL FORMAT
            converters = {
                    'aircraft': (lambda x: 'LN'+x.strip())
                    }
            df = pd.read_excel(fname, converters=converters)
            df['Workordernumber'] = df['Workordernumber'].map(str).str.strip()
            
            conc = 'oracle://rcp_con:CDF34@ora-pr06flm:1527/pr06'
            engine = create_engine(conc)

            #GETTING THE DATE REF
            date_ref = preferences['date_entry_upload_since']
            date_ref = pd.to_datetime(date_ref, format='%d/%m/%Y')

            SQL = Return_SQL()
            print('EXECTUTANDO A CONSLUTA DE INTER PARA A WIDEROE')
            dfquery = pd.read_sql_query(SQL, engine, params={'DATE_REF': date_ref.strftime("%d/%m/%Y")})
            print('CONSULTA EXECUTADA')
            df_wideroe = pd.DataFrame(dfquery)
            df_wideroe['Fisca'] = df_wideroe['Fisca'].map(str).str.strip()

            df_inter = pd.merge(df, df_wideroe, left_on=['Workordernumber'], 
                                right_on=['Fisca'], how='inner')
            df_inter = df_inter.filter(items = ['Workordernumber', 'Workorder_Desc_text', 'Workorder_Action_text', 'Event Description', 'Action Description', 'Problem'])
            df_inter = df_inter.rename({'Event Description':'DESCRIPTION_INTER', 
                                                      'Action Description':'ACTION_INTER',
                                                      'Workorder_Desc_text':'DESCRIPTION_OOS',
                                                      'Workorder_Action_text':'ACTION_OOS',
                                                      'Workordernumber':'LOGNUMBER'}, axis=1)
            
            with pd.ExcelWriter("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/WIDEROE/INTER_WIDEROE.xlsx", engine="xlsxwriter") as writer:
                df_inter.to_excel(writer, sheet_name="INTER", index=False)
            writer.save()

            #FORMATING THE WORKSHEET
            print('Fromatando as abas')
            excel = Dispatch('Excel.Application')
            wb_formated = excel.Workbooks.Open("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/WIDEROE/INTER_WIDEROE.xlsx")
            excel.Application.ScreenUpdating = False
            excel.Worksheets(1).Activate()
            excel.ActiveSheet.Columns("A").AutoFit()
            excel.ActiveSheet.Columns("B").ColumnWidth = 19
            excel.ActiveSheet.Columns("C").ColumnWidth = 19
            excel.ActiveSheet.Columns("D").ColumnWidth = 19
            excel.ActiveSheet.Columns("E").ColumnWidth = 19
            excel.ActiveSheet.Columns("F").AutoFit
            excel.Application.ScreenUpdating = True
            wb_formated.Save()
            wb_formated.Close()
            print('Formatação concluída!')
            
            print('Arbindo as planilhas')
            wb = xw.Book("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/WIDEROE/INTER_WIDEROE.xlsx")
            print('Finalizado!')

        get_cleaned_df(self, fname)

    @staticmethod    
    def merge_helvetic(self):
        print('Executando o relacionamento para a HELVETIC AIRWAYS...')
        from datetime import date, timedelta
        warnings.filterwarnings('ignore')
        def get_fname(preferences):

            ano = preferences["date_entry_upload_since"].split('/')[2]
            mes = preferences["date_entry_upload_since"].split('/')[1]
            fname = '//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/1 - Dados recebidos/4 - EMEA/HELVETIC AIRWAYS' + '/' + ano + '/' + mes + '/' + 'E2/' + 'OOS_DATA-HELVETIC.xlsx'
            return fname

        #DEFINE THE DIRECTORY OF THE FILE
        preferences = read_preferences()
        fname = get_fname(preferences)
        
        def Return_SQL():
            
            return ("""
                SELECT
                    COD_SEQ_PROB AS "Problem",   
                    REG_AENV_DATA_EVNT AS "Register",
                    NOM_SIMPL_ORGZ_DATA_EVNT AS "Operator (Simplified Name)",
                    COD_CAP_ATA_FALHA  AS "Ata Chapter",   
                    COD_SUCP_ATA_FALHA AS "Ata Sub. Chapter",
                    NUM_FICSA AS "Fisca", 
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 
                    'PROB:', 1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 
                    'ACTION:', 1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 1, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) -1)  ELSE  DSC_ACAO_CRRT  END AS "Event Description", 
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 'PROB:', 
                    1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1))  ELSE  DSC_ACAO_CRRT  END AS "Action Description",
                    to_char(DAT_EVNT,'MM/dd/yyyy') AS "Event Date"
                FROM
                    RCP_VW_PROB_AERONAVE wp 
                    WHERE 1 = 1   AND DAT_REF >= TO_DATE(:DATE_REF,'DD/MM/YYYY')  
                    AND DAT_REF <= TO_DATE(:DATE_REF,'DD/MM/YYYY')
                    AND COD_TIPO_PROB = 'INTER'
                    AND COD_MERC_AVC = 'C'   AND COD_PROJ IN ('1902','1952')   
                    AND COD_ORGZ_DATA_EVNT = 10951
            """)
            
        def get_cleaned_df(self, fname):
            preferences = read_preferences()
            fname = get_fname(preferences)
            
            converters = {
                    'Event A/C': (lambda x: 'HB-'+x.strip())
                    }
            #READING AND FORMATING THE OOS DATA
            df = pd.read_excel(fname, converters=converters)
            df['Workorder Number'] = df['Workorder Number'].map(str).str.strip()
            df['Event A/C'] = df['Event A/C'].map(str).str.strip()
            df['Occurrence Date'] = pd.to_datetime(df['Occurrence Date'].map(str), format='%Y/%m/%d')
            df['Occurrence Date'] = df['Occurrence Date'].dt.strftime('%d/%m/%Y')
            df['Occurrence Date'] = df['Occurrence Date'].map(str)
            df['Occurrence Date'] = df['Occurrence Date'].str.strip()
            df['ATA Chapter'] = df['ATA Chapter'].map(str)
            df['ATA Chapter'] = df['ATA Chapter'].str.strip()
            df['ATA Chapter'] = df['ATA Chapter'].str.lstrip('0')
            '''
            #SEPARATING ATA AND SUB-ATA
            ata = df['ATA Chapter'].str.split('-')
            if ata[0] == null: 
                df['ATA Chapter'] =  ata[1]
            else:
                df['ATA Chapter'] =  ata[0]
            '''

            #Fill the NULL fields
            df.fillna('', inplace=True)
            
            conc = 'oracle://rcp_con:CDF34@ora-pr06flm:1527/pr06'
            engine = create_engine(conc)
            date_ref = preferences['date_entry_upload_since']
            date_ref = pd.to_datetime(date_ref, format='%d/%m/%Y')
            SQL = Return_SQL()
            print('RUNNING THE QUERY')
            dfquery = pd.read_sql_query(SQL, engine, params={'DATE_REF': date_ref.strftime("%d/%m/%Y")})
            print('QUERY EXECUTED')
            df_helvetic_inter = pd.DataFrame(dfquery)
            df_helvetic_inter['Fisca'] = df_helvetic_inter['Fisca'].map(str).str.strip()
            df_helvetic_inter['Event Date'] = pd.to_datetime(df_helvetic_inter['Event Date'].map(str), format='%m/%d/%Y')
            df_helvetic_inter['Event Date'] = df_helvetic_inter['Event Date'].dt.strftime('%d/%m/%Y')
            df_helvetic_inter['Event Date'] = df_helvetic_inter['Event Date'].map(str)
            df_helvetic_inter['Event Date'] = df_helvetic_inter['Event Date'].str.strip()
            df_helvetic_inter['Ata Chapter'] = df_helvetic_inter['Ata Chapter'].map(str)
            df_helvetic_inter['Ata Chapter'] = df_helvetic_inter['Ata Chapter'].str.strip()
            df_helvetic_inter.to_excel('DF_HELVETIC.xlsx', sheet_name='DF_HELVETIC', index=False)

            df_inter = pd.merge(df, df_helvetic_inter, left_on=['Workorder Number'], 
                                right_on=['Fisca'], how='inner')
            df_inter = df_inter.filter(items = ['Problem', 'Occurrence Date', 'Event A/C', 'ATA Chapter', 'Workorder_Desc_text', 'Workorder_Action_text', 'Fisca', 'Workorder Number', 'Workorder Text', 'Workorder Action'])
            if df['Workorder Number'].any:
                df_inter_without_wo = pd.merge(df, df_helvetic_inter, left_on=['Occurrence Date', 'Event A/C', 'ATA Chapter'], 
                                right_on=['Event Date','Register', 'Ata Chapter'], how='inner')
                #df_inter_without_wo = df_inter_without_wo[df_inter_without_wo['Fisca'] == None]
                #df_inter_without_wo = df_inter_without_wo.filter(items = ['ATA Chapter', 'Occurrence Date', 'Event A/C', 'Workorder Number', 'Workorder Text', 'Workorder Action', 'Problem', 'Fisca', 'Event Description', 'Action Description'])
                df_inter_without_wo = df_inter_without_wo.filter(items = ['Problem', 'Occurrence Date', 'Event A/C', 'ATA Chapter', 'Event Description', 'Action Description', 'Fisca',  'Workorder Number', 'Workorder Text', 'Workorder Action'])
                df_inter_without_wo.to_excel("df_inter_without_wo.xlsx", sheet_name='df_inter_without_wo')
            df_inter_full = pd.concat([df_inter, df_inter_without_wo], ignore_index=True)

            df_inter_full = df_inter_full.filter(items = ['Problem', 'Occurrence Date', 'Event A/C', 'ATA Chapter', 'Workorder Number', 'Event Description', 'Action Description', 'Workorder Text', 'Workorder Action'])
            df_inter_full = df_inter_full.rename({'Event Description':'DESCRIPTION_INTER', 
                                                      'Action Description':'ACTION_INTER',
                                                      'Workorder Text':'DESCRIPTION_OOS',
                                                      'Workorder Action':'ACTION_OOS',
                                                      'Occurrence Date':'DATE',
                                                      'ATA Chapter':'ATA',
                                                      'Event A/C':'REGISTER',
                                                      'Workorder Number':'LOGNUMBER_OOS'}, axis=1)
                                                    
            
            with pd.ExcelWriter("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/HELVETIC/INTER_HELVETIC.xlsx", engine="xlsxwriter") as writer:
                df_inter_full .to_excel(writer, sheet_name="INTER", index=False)
            writer.save()

            print('Fromatando as abas')
            excel = Dispatch('Excel.Application')
            wb_formated = excel.Workbooks.Open("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/HELVETIC/INTER_HELVETIC.xlsx")
            excel.Application.ScreenUpdating = False
            excel.Worksheets(1).Activate()
            excel.ActiveSheet.Columns("A:E").AutoFit()
            excel.ActiveSheet.Columns("F").ColumnWidth = 19
            excel.ActiveSheet.Columns("G").ColumnWidth = 19
            excel.ActiveSheet.Columns("H").ColumnWidth = 19
            excel.ActiveSheet.Columns("I").ColumnWidth = 19
            excel.Application.ScreenUpdating = True
            wb_formated.Save()
            wb_formated.Close()
            print('Formatação concluída!')
            
            print('Arbindo as planilhas')
            wb = xw.Book("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/HELVETIC/INTER_HELVETIC.xlsx")
            print('Finalizado!')

        get_cleaned_df(self, fname)

    @staticmethod
    def merge_azul(self):
        from datetime import date, timedelta
        print("Executando o relacionamento da AZUL...")
        warnings.filterwarnings('ignore')
        def get_fname(preferences):

            ano = preferences["date_entry_upload_since"].split('/')[2]
            mes = preferences["date_entry_upload_since"].split('/')[1]
            fname = '//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/1 - Dados recebidos/5 - LATIN AMERICA/AZUL/' + ano + '/' + mes + '/' + 'OOS_DATA-AZUL.xlsx'
            return fname

        def _skip_blank_header(df):
            df = df.dropna(how='all').reset_index(drop=True)
            while sum([not (pd.isna(col) or 'Unnamed' in col) for col in df.columns]) < 5:
                df.columns = df.iloc[0]
                df = df.drop([0], axis=0).reset_index(drop=True)
        
            return df        
        
        def load_file_azul_pirep(self, preferences, fname, converters=None, dtype=None, sheet_name=None):
            fname = get_fname(preferences)            
            df = pd.read_excel(fname, converters=converters, dtype=dtype, sheet_name='LOGBOOKS')
            df = _skip_blank_header(df)
            df.columns = [col.strip().lower() for col in df.columns]
            return df

        def load_file_azul(self, preferences, fname, converters=None, dtype=None, sheet_name=None):
            fname = get_fname(preferences)
            df = pd.read_excel(fname, converters=converters, dtype=dtype, sheet_name='OOS')
            df = _skip_blank_header(df)
            df.columns = [col.strip().lower() for col in df.columns]
            return df
        
        #DEFINE THE DIRECTORY OF THE FILE
        preferences = read_preferences()
        fname = get_fname(preferences)
        
        def Return_SQL():
            return ("""
                SELECT
                    COD_SEQ_PROB AS "Problem",   
                    REG_AENV_DATA_EVNT AS "Register",
                    NOM_SIMPL_ORGZ_DATA_EVNT AS "Operator (Simplified Name)",
                    COD_CAP_ATA_FALHA  AS "Ata Chapter",   
                    COD_SUCP_ATA_FALHA AS "Ata Sub. Chapter",  
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 
                    'PROB:', 1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 
                    'ACTION:', 1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 1, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) -1)  ELSE  DSC_ACAO_CRRT  END AS "Event Description", 
                    CASE  WHEN INSTR(DSC_ACAO_CRRT, 'PROB:', 
                    1) <> 0 AND  INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1) <> 0  THEN SUBSTR(DSC_ACAO_CRRT, 
                    INSTR(DSC_ACAO_CRRT, 'ACTION:', 
                    1))  ELSE  DSC_ACAO_CRRT  END AS "Action Description",
                    to_char(DAT_EVNT,'MM/dd/yyyy') AS "Event Date"
                FROM
                    RCP_VW_PROB_AERONAVE wp 
                    WHERE 1 = 1   AND DAT_REF >= TO_DATE(:DATE_REF,'DD/MM/YYYY')  
                    AND DAT_REF <= TO_DATE(:DATE_REF,'DD/MM/YYYY')
                    AND COD_TIPO_PROB = 'INTER'
                    AND COD_MERC_AVC = 'C'   AND COD_PROJ IN ('1902','1952')   
                    AND COD_ORGZ_DATA_EVNT = 10040

            """)
                    
        def get_cleaned_df(self, fname):
            preferences = read_preferences()
            fname = get_fname(preferences)
            
            df = load_file_azul(self, preferences, fname)
            #df = pd.read_excel(fname, sheet_name='OOS', skiprows=1)
            df_filter = pd.DataFrame(df)
            df_filter['chapter'] = df_filter['chapter'].map(str)
            #SEPARATING THE INICIO AND TERMINO COLUMNS IN data_inicio, hora_inicio, data_termino, hora_termino
            df_filter['inicio'] = df_filter['inicio'].map(str).str.strip()
            data_hora_inicio = df_filter['inicio'].str.split(' ', n=1, expand=True)
            df_filter['data_inicio'] = data_hora_inicio[0]
            df_filter['data_inicio'] = pd.to_datetime(df_filter['data_inicio'], format='%Y/%m/%d')
            df_filter['data_inicio'] = df_filter['data_inicio'].dt.strftime('%d/%m/%Y')

            df_filter_proc =  df_filter.filter(items = ['ac', 'chapter', 'defect_description', 
                                                        'resolution_description'])
            df_filter_proc['reported date'] = pd.to_datetime(df_filter['data_inicio'], format='%d/%m/%Y')
            df_filter_proc['reported date'] = df_filter_proc['reported date'].dt.strftime('%d/%m/%Y')
            df_filter_proc['reported date'] = df_filter_proc['reported date'].map(str)
            df_filter_proc['reported date'] = df_filter_proc['reported date'].str.strip()

            df_pirep_filter = load_file_azul_pirep(self, preferences, fname)
            df_pirep_filter['ac'] = df_pirep_filter['ac'].str.strip()
            df_pirep_filter['ac_x'] = df_pirep_filter['ac'].str.strip() 
            df_pirep_filter['ata'] = df_pirep_filter['ata'].map(str)
            #SEPARATING THE INICIO AND TERMINO COLUMNS IN data_inicio, hora_inicio, data_termino, hora_termino
            df_pirep_filter['reported date'] = df_pirep_filter['reported date'].map(str).str.strip()
            data_hora_inicio = df_pirep_filter['reported date'].str.split(' ', n=1, expand=True)
            df_pirep_filter['reported date'] = data_hora_inicio[0]
            df_pirep_filter['reported date'] = pd.to_datetime(df_pirep_filter['reported date'], format='%d/%m/%Y')
            df_pirep_filter['reported date'] = df_pirep_filter['reported date'].dt.strftime('%d/%m/%Y')
            '''
            df_pirep_filter['reported date'] = df_pirep_filter['reported date'].dt.strftime('%d/%m/%Y')
            df_pirep_filter['reported date'] = df_pirep_filter['reported date'].map(str)
            df_pirep_filter['reported date'] = df_pirep_filter['reported date'].str.strip()
            '''
            
            conc = 'oracle://rcp_con:CDF34@ora-pr06flm:1527/pr06'
            engine = create_engine(conc)

            #GET DATE OF REFERENCE
            date_ref = preferences['date_entry_upload_since']
            date_ref = pd.to_datetime(date_ref, format='%d/%m/%Y')
            SQL = Return_SQL()
            print('RUNNING THE QUERY')
            dfquery = pd.read_sql_query(SQL, engine, params={'DATE_REF': date_ref.strftime("%d/%m/%Y")})
            print('QUERY EXECUTED')
            df_azul = pd.DataFrame(dfquery)
            df_azul['Event Date'] = pd.to_datetime(df_azul['Event Date'].map(str), format='%m/%d/%Y')
            df_azul['Event Date'] = df_azul['Event Date'].dt.strftime('%d/%m/%Y')
            df_azul['Event Date'] = df_azul['Event Date'].map(str)
            df_azul['Event Date'] = df_azul['Event Date'].str.strip()
            df_azul['Ata Chapter'] = df_azul['Ata Chapter'].map(str)
            df_azul['Ata Chapter'] = df_azul['Ata Chapter'].str.strip()
            df_filter_inter = pd.merge(df_filter_proc, df_azul, left_on=['reported date','ac', 'chapter'], 
                                       right_on=['Event Date','Register', 'Ata Chapter'], 
                                       how='inner')
            df_filter_inter = df_filter_inter.filter(items = ['Problem', 'reported date', 'ac', 'chapter',
                                                              'defect_description', 'Event Description',
                                                              'resolution_description', 'Action Description'])
            df_filter_inter = df_filter_inter.rename({'Event Description':'DESCRIPTION_INTER', 
                                                      'Action Description':'ACTION_INTER',
                                                      'defect_description':'DESCRIPTION_OOS',
                                                      'resolution_description':'ACTION_OOS',
                                                      'reported date':'DATE',
                                                      'chapter':'ATA',
                                                      'ac':'REGISTER'}, axis=1)
               
            df_filter_pirep = pd.merge(df_filter_proc, df_pirep_filter, left_on=['reported date', 'ac', 'chapter'], 
                                           right_on=['reported date', 'ac', 'ata'], 
                                           how='inner')
            
            df_filter_pirep = df_filter_pirep.filter(items = ['reported date', 
                                                               'ac', 'chapter',
                                                               'defect_description_x', 'defect_description_y',
                                                               'resolution_description_x', 
                                                               'resolution_description_y'])
            df_filter_pirep = df_filter_pirep.rename({'defect_description_y':'DESCRIPTION_PIREP', 
                                                      'resolution_description_y':'ACTION_PIREP',
                                                      'defect_description_x':'DESCRIPTION_OOS',
                                                      'resolution_description_x':'ACTION_OOS',
                                                      'reported date':'DATE',
                                                      'chapter':'ATA',
                                                      'ac':'REGISTER'}, axis=1)
            
            with pd.ExcelWriter("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/AZUL/PIREP_INTER_AZUL.xlsx", engine="xlsxwriter") as writer:
                df_filter_pirep.to_excel(writer, sheet_name="PIREP", index=False)
                df_filter_inter.to_excel(writer, sheet_name="INTER", index=False)
            writer.save()

            print('Fromatando as abas')
            excel = Dispatch('Excel.Application')
            wb_formated = excel.Workbooks.Open("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/AZUL/PIREP_INTER_AZUL.xlsx")
            excel.Application.ScreenUpdating = False
            excel.Worksheets(1).Activate()
            excel.ActiveSheet.Columns("A:C").AutoFit()
            excel.ActiveSheet.Columns("B").NumberFormat = 'DD/MM/AAAA'
            excel.ActiveSheet.Columns("D").ColumnWidth = 19
            excel.ActiveSheet.Columns("E").ColumnWidth = 19
            excel.ActiveSheet.Columns("F").ColumnWidth = 19
            excel.ActiveSheet.Columns("G").ColumnWidth = 19
            excel.Worksheets(2).Activate()
            excel.ActiveSheet.Columns("A:D").AutoFit()
            excel.ActiveSheet.Columns("E").ColumnWidth = 19
            excel.ActiveSheet.Columns("F").ColumnWidth = 19
            excel.ActiveSheet.Columns("G").ColumnWidth = 19
            excel.ActiveSheet.Columns("H").ColumnWidth = 19
            excel.Application.ScreenUpdating = True
            wb_formated.Save()
            wb_formated.Close()
            print('Formatação concluída!')
            
            print('Arbindo as planilhas')
            wb = xw.Book("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/AZUL/PIREP_INTER_AZUL.xlsx")
            print('Finalizado!')
        get_cleaned_df(self, fname)

class AzulParser(Parser):
    #Get the AZUL path
    def __init__(self, root, file_pattern):
        self.root = root
        self.azul_path = '/5 - LATIN AMERICA/AZUL'
        super(AzulParser, self).__init__(
            self.root + self.azul_path, file_pattern)
        
    #Get all spreadsheet records
    def get_cleaned_df(self, fname):

        warnings.filterwarnings('ignore')
        #Cleaning the screen
        os.system('cls')
        #DATA PROCESSING
        df = self.load_file_azul(fname)
        df_filter = pd.DataFrame(df)
        #df_filter['fleet'] = df_filter['fleet'].str.strip()
        #df_filter['fleet'] = df_filter[df_filter['fleet'] == 'E2']
        df_filter['station'] = df_filter['station'].str.rstrip('-H')
        df_filter['chapter'] = df_filter['chapter'].map(str).str.strip()
        
        #SEPARATING THE INICIO AND TERMINO COLUMNS IN data_inicio, hora_inicio, data_termino, hora_termino
        df_filter['inicio'] = df_filter['inicio'].map(str).str.strip()
        df_filter['termino'] = df_filter['termino'].map(str).str.strip()
        data_hora_inicio = df_filter['inicio'].str.split(' ', n=1, expand=True)
        data_hora_termino = df_filter['termino'].str.split(' ', n=1, expand=True)
        df_filter['data_inicio'] = data_hora_inicio[0]
        df_filter['hora_inicio'] = data_hora_inicio[1]
        df_filter['data_termino'] = data_hora_termino[0]
        df_filter['hora_termino'] = data_hora_termino[1]
        df_filter['data_inicio'] = pd.to_datetime(df_filter['data_inicio'], format='%Y/%m/%d')
        df_filter['data_inicio'] = df_filter['data_inicio'].dt.strftime('%d/%m/%Y')
        df_filter['data_termino'] = pd.to_datetime(df_filter['data_termino'], format='%Y/%m/%d')
        df_filter['data_termino'] = df_filter['data_termino'].dt.strftime('%d/%m/%Y')
        
        #df_filter['event_description'] = df_filter['defect_description']
        #df_filter['action_description'] = df_filter['resolution_description']
        

        #CRIATION OF DATASETS INTER e PIREP
        print('Lendo os dados de Inter e Pirep da Azul...')
        df_pirep_filter = pd.read_excel("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/AZUL/PIREP_INTER_AZUL.xlsx", sheet_name='PIREP')
        df_inter_filter = pd.read_excel("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/AZUL/PIREP_INTER_AZUL.xlsx", sheet_name='INTER')
        print('Finalizado')

        df_pirep_filter['ATA'] = df_pirep_filter['ATA'].map(str).str.strip()
        df_inter_filter['ATA'] = df_pirep_filter['ATA'].map(str).str.strip()
        df_inter_filter['DATE'] = pd.to_datetime(df_inter_filter['DATE'], format='%d/%m/%Y')
        df_inter_filter['DATE'] = df_inter_filter['DATE'].dt.strftime('%d/%m/%Y')
        
        #Cleaning the screen    
        os.system('cls')
        print("CONCATENANDO OS ARQUIVOS DA AZUL")
        #Concatenando o dataframe original com o de PIREP
        df_filter = pd.merge(df_pirep_filter, df_filter, right_on=['data_inicio', 'chapter', 'ac', 'defect_description'], left_on=['DATE', 'ATA', 'REGISTER', 'DESCRIPTION_OOS'], how='right')
        #Concatenando o dataframe original com o de INTER

        df_filter = pd.merge(df_inter_filter, df_filter, right_on=['data_inicio', 'chapter', 'ac', 'defect_description'], left_on=['DATE', 'ATA', 'REGISTER', 'DESCRIPTION_OOS'], how='right')

        '''
        #Retira as colunas duplicadas após o merge
        df_filter.drop('resolution_description_y', axis=1, inplace=True)
        df_filter.drop('defect_description_y', axis=1, inplace=True)
        df_filter.drop('resolution_description_x', axis=1, inplace=True)
        df_filter.drop('defect_description_x', axis=1, inplace=True)
        '''

        df_filter = df_filter.filter(items=['Problem', 'ac', 'status', 'data_inicio', 'hora_inicio', 'data_termino', 'hora_termino', 'time_duration', 'station', 'defect', 'chapter', 'defect_description', 'resolution_description', 'ACTION_PIREP'])
        df_filter.fillna('', inplace=True)
        #df_filter['action_description'] =  df_filter['action_description'].map(str)
        #df_filter['ACTION_PIREP'].map(str)
        df_filter['defect_description'] = df_filter['defect_description'].map(str)
        df_filter['resolution_description'] = df_filter['resolution_description'].map(str)
        df_filter['resolution_description_aux'] = '**PIREP ACTION: ' + df_filter['ACTION_PIREP'].map(str) + ' **'
        df_filter['resolution_description'] = df_filter['resolution_description_aux'] + '\n' + df_filter['resolution_description'].map(str)
        df_filter['resolution_description'] = df_filter['resolution_description'].map(str)
        #df_filter.drop(columns=['inicio, termino'], inplace=True)
        print("FINALIZADO")
        cleaned_df = pd.DataFrame()
        try:
            print("INICIANDO A PREPARAÇÂO PARA A IMPORTAÇÂO DOS DADOS")
            cleaned_df['Aircraft_Register__c'] = df_filter['ac'].str.strip()
            cleaned_df['Start_Date__c'] = self.normalize_datetime(
                df_filter, 'data_inicio', 'hora_inicio')
            cleaned_df['Release_Date__c'] = self.normalize_datetime(
                df_filter, 'data_termino', 'hora_termino')
    
            df_filter.fillna('', inplace=True)
    
            cleaned_df['Log_Number__c'] = df_filter['defect'].astype(str)
            cleaned_df['OOS_Total_Time__c'] = self.normalize_oos_time(df_filter, cleaned_df, 'time_duration')
            cleaned_df['Station__c'] = df_filter['station'].str.strip()
            cleaned_df['Operator_ATA_Chapter__c'] = df_filter['chapter']
            cleaned_df['Event_Record_Identifier__c'] = df_filter['status'].fillna(
                '').str.strip()
            cleaned_df['Event_Description__c'] = df_filter['defect_description']
            cleaned_df['Action_Description__c'] = df_filter['resolution_description']
            cleaned_df['Inter_ID__c'] = df_filter['Problem']
            cleaned_df['Reference_Date__c'] = self.get_reference_date(
                fname, len(cleaned_df))
            
            cleaned_df.to_excel('CLEANED_DF_AZUL.xlsx', sheet_name='df', index=False)
            cleaned_df = self._commom_clean(cleaned_df)
            #os.remove(fname)
            print("FINALIZADO")
        except KeyError as e:
            error = 'FAIL - KEYERROR - NOT RECOGNIZE THE USUAL COLUMN ' + str(e)
            return error, cleaned_df
            
        return "OK", cleaned_df

class WideroeParser(Parser):

    def __init__(self, root, file_pattern):
        self.root = root
        self.wideroe_path = '/4 - EMEA/WIDEROE'
        super(WideroeParser, self).__init__(
            self.root + self.wideroe_path, file_pattern)
                    
    def get_cleaned_df(self, fname):
        warnings.filterwarnings('ignore')
        converters = {
            'aircraft': (lambda x: x if x.startswith('LN') else 'LN-' + x.strip())}
        
        #READING THE WIDEROE OOS DATA
        df = self.load_file(fname, converters=converters)
        df_wideroe_inter = pd.read_excel("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/WIDEROE/INTER_WIDEROE.xlsx")
        df['workordernumber'] = df['workordernumber'].map(str).str.strip()
        df_wideroe_inter['LOGNUMBER'] = df_wideroe_inter['LOGNUMBER'].map(str).str.strip()   
        df = pd.merge(df, df_wideroe_inter, left_on=['workordernumber'], 
                                right_on=['LOGNUMBER'], how='left')
        #df = df.filter(items = ['Problem', 'DESCRIPTION_OOS', 'ACTION_OOS', 'LOGNUMBER'])
        df = df.filter(items = ['Problem', 'aircraft', 'oos_start_date_and_time', 'oos_end_date_and_time', 'workordernumber', 'oos_total_hrs_downtime', 'workorder_ata', 'ops_code', 'workorder_desc_text', 'workorder_action_text', 'event_header', 'flightnumber', 'station'])   
        df['aircraft'] = np.where(df['aircraft'].str.len() < 4, np.nan, df['aircraft'])
        df.dropna(subset=['aircraft'], inplace=True)
        df['Problem'] = df['Problem'].map(str)
        df['Problem'] = df['Problem'].map(lambda x: str(x).replace('.', '').rstrip('0'))
        df.to_excel("DF.xlsx", sheet_name='DF')
        try:
            cleaned_df = pd.DataFrame()
            df['aircraft'] = np.where(df['aircraft'].str.len() < 4, np.nan, df['aircraft'])
            df.dropna(subset=['aircraft'], inplace=True)
            cleaned_df['Aircraft_Register__c'] = df['aircraft']
            cleaned_df['Start_Date__c'] = self.normalize_datetime(
                df, 'oos_start_date_and_time')
            cleaned_df['Release_Date__c'] = self.normalize_datetime(
                df, 'oos_end_date_and_time')
    
            df.fillna('', inplace=True)
    
            cleaned_df['Log_Number__c'] = df['workordernumber']
            cleaned_df['OOS_Total_Time__c'] = self.normalize_oos_time(df, cleaned_df, 'oos_total_hrs_downtime')
            cleaned_df['Station__c'] = df['station'].str.strip()
            cleaned_df['Operator_ATA_Chapter__c'] = df['workorder_ata']
            cleaned_df['Event_Record_Identifier__c'] = df['ops_code'].str.strip()
            cleaned_df['Event_Description__c'] = df['workorder_desc_text']
            cleaned_df['Action_Description__c'] = df['workorder_action_text']
            cleaned_df['Inter_ID__c'] = df['Problem']
            cleaned_df['Header__c'] = df['event_header']
            cleaned_df['Flight_Number__c'] = df['flightnumber'].astype(str)
            cleaned_df['Reference_Date__c'] = self.get_reference_date(
                fname, len(cleaned_df))
            
            cleaned_df = self._commom_clean(cleaned_df)
            cleaned_df.fillna('', inplace=True)
            os.remove(fname)
            os.remove("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/WIDEROE/INTER_WIDEROE.xlsx")
        except KeyError as e:
            error = 'FAIL - KEYERROR - NOT RECOGNIZE THE USUAL COLUMN ' + str(e)
            return error, cleaned_df
        
        return "OK", cleaned_df

class HelveticParser(Parser):

    def __init__(self, root, file_pattern):
        self.root = root
        self.helvetic_path = '/4 - EMEA/HELVETIC AIRWAYS'
        super(HelveticParser, self).__init__(
            self.root + self.helvetic_path, file_pattern)

    @staticmethod
    def _get_event_description(s):
        if not pd.isna(s['workorder text']) and len(s['workorder text'].strip()) > 0:
            return s['workorder text']
        else:
            return s['header']
        #re.findall(r'(?:(?:Technical)|(?:AOG))\s*Event\:\s*([\s\S]+?)\;', s['description'], re.IGNORECASE)[0]

    @staticmethod
    def _get_event_action(s):
        if not pd.isna(s['workorder action']) and len(s['workorder action'].strip()) > 0:
            return s['workorder action']
        else:
            return s['event data']
            #return re.findall(r'Solution\:\s*([\s\S]+?)\;', s['description'], re.IGNORECASE)[0]

    def get_cleaned_df(self, fname):
        warnings.filterwarnings('ignore')
        converters = {
            'occurrence date': self._date_format_f, 
            'ready date': self._date_format_f,
            'Event A/C': (lambda x: 'HB-'+x.strip())}
        df = self.load_file(fname, converters=converters)
        df['workorder number'] = df['workorder number'].map(str).str.strip()
        df['event a/c'] = df['event a/c'].map(str).str.strip()
        df['occurrence date'] = pd.to_datetime(df['occurrence date'].map(str), format='%Y/%m/%d')
        df['occurrence date'] = df['occurrence date'].dt.strftime('%d/%m/%Y')
        df['occurrence date'] = df['occurrence date'].map(str)
        df['occurrence date'] = df['occurrence date'].str.strip()
        df['ata chapter'] = df['ata chapter'].map(str)
        df['ata chapter'] = df['ata chapter'].str.strip()
        df['ata chapter'] = df['ata chapter'].str.lstrip('0')


        df_helvetic = pd.read_excel("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/HELVETIC/INTER_HELVETIC.xlsx")
        df_helvetic['LOGNUMBER_OOS'] = df_helvetic['LOGNUMBER_OOS'].map(str).str.strip()
        df_inter = pd.merge(df, df_helvetic, left_on=['workorder number', 'occurrence date', 'event a/c', 'ata chapter', 'workorder text'], 
                                right_on=['LOGNUMBER_OOS', 'DATE','REGISTER', 'ATA', 'DESCRIPTION_OOS'], how='left')
        df.fillna('', inplace=True)
        df_inter['workorder number'] = df_inter['workorder number'].astype(int)
        df =  df_inter.filter(items=['event a/c', 'workorder text', 'header', 'workorder action', 'description', 'occurrence date', 'occurrence time', 'ready date', 'ready time', 'workorder number', 'Problem', 'repair station', 'ata chapter', 'event flight number'])
        df.dropna(subset=['event a/c'], inplace=True)
        df.to_excel('DF_HELVETIC.xlsx', sheet_name='DF', index=False)
        df = pd.DataFrame(df)

        cleaned_df = pd.DataFrame()
        try:
            

            cleaned_df['Aircraft_Register__c'] = df['event a/c'].str.strip()
            #cleaned_df['Event_Record_Identifier__c'] = df['description'].apply(str).str.extract(
                #r'Status\:\s*(.*?)\s*\;', re.IGNORECASE, expand=False).str.strip()
            cleaned_df['Event_Description__c'] = df[['workorder text',
                                                     'header']].apply(self._get_event_description, axis=1)
            cleaned_df['Action_Description__c'] = df[['workorder action',
                                                      'description']].apply(self._get_event_action, axis=1)
            cleaned_df['Header__c'] = df['header']
                
            
            cleaned_df['Start_Date__c'] = self.normalize_datetime(
                df, 'occurrence date', 'occurrence time')
            cleaned_df['Release_Date__c'] = self.normalize_datetime(
                df, 'ready date', 'ready time')
    
            
    
            cleaned_df['OOS_Total_Time__c'] = self.normalize_oos_time(df, cleaned_df, None)
            
            cleaned_df['Log_Number__c'] = df['workorder number'].map(
                int).apply(lambda x: np.nan if x <= 0 else x)
            cleaned_df['Inter_ID__c'] = df['Problem']
            cleaned_df['Station__c'] = df['repair station'].str.strip()
            cleaned_df['Operator_ATA_Chapter__c'] = df['ata chapter']
            
            cleaned_df['Flight_Number__c'] = df['event flight number'].astype(str)
            cleaned_df['Reference_Date__c'] = self.get_reference_date(
                fname, len(cleaned_df))
            
            cleaned_df = self._commom_clean(cleaned_df)
            cleaned_df.fillna('', inplace=True)
            cleaned_df.to_excel('CLEANED_DF.xlsx', sheet_name='DF', index=False)
            #os.remove(fname)
            #os.remove('//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2//INTER_PIREP/HELVETIC/INTER_HELVETIC.xlsx')
        except KeyError as e:
            error = 'FAIL - KEYERROR - NOT RECOGNIZE THE USUAL COLUMN ' + str(e)
            return error, cleaned_df
        
        return "OK", cleaned_df
class AstanaParser(Parser):

    def __init__(self, root, file_pattern):
        self.root = root
        self.astana_path = '/4 - EMEA/AIR ASTANA'
        super(AstanaParser, self).__init__(
            self.root + self.astana_path, file_pattern)

    @staticmethod
    def _correlate_columns(df):
        c_correlated = {}
        for col in df.columns:
            if col == 'other':
                c_correlated[col] = 'Others__c'
            elif col == 'parts unavailability':
                c_correlated[col] = 'Parts_Unavailability__c'
            elif 'receive embraer disposition' in col:
                c_correlated[col] = 'Time_to_Receive_Embraer_Disposition__c'
            elif 'customer operations' in col:
                c_correlated[col] = 'Customer_Operation__c'
            elif 'time for troubleshooting' in col:
                c_correlated[col] = 'Time_to_Receive_Supplier_Disposition__c'

        return c_correlated

    def _prepare_df(self, df):
        # while df.columns[0] != 'A/C':
        #     if df.iloc[0, 0] == 'A/C':
        #         col_names = df.iloc[0]
        #         df.columns = col_names
        #     df = df.drop([0], axis=0).reset_index(drop=True)
            
        categories = df[['category', 'contrib (%)', 'comments']].copy()
        categories['a/c'] = df['a/c'].ffill().values
        categories['contrib (%)'].fillna(0, inplace=True)
        categories_pivot = categories.pivot_table('contrib (%)', ['a/c'], 'category')
        
        ngroups = categories.groupby('category').ngroups
        ac, comments = np.asarray(np.split(df[['a/c', 'comments']].fillna('').values, ngroups)).T
        
        ac = [''.join(cell for cell in row) for row in ac]
        comments = [''.join(cell for cell in row) for row in comments]
        
        categories_pivot = pd.merge(categories_pivot, 
                                    pd.DataFrame({'comments': comments, 'a/c': ac}),
                                    on='a/c')
        
        df = df.replace(r'^\s*$', np.nan, regex=True)
        df = df.dropna(axis='rows', subset=[
            'a/c']).drop(columns=['category', 'contrib (%)', 'comments'])
        
        df = pd.merge(df, categories_pivot, on='a/c')
        df.columns = [col.strip().lower() for col in df.columns]

        correlated_columns = self._correlate_columns(df)
        for techrep_col in correlated_columns.keys():
            df[techrep_col] = df[techrep_col].apply(
                lambda x: x*100 if x <= 1 else x).fillna('').astype(str)

        return df, correlated_columns

    def get_cleaned_df(self, fname):

        df = self.load_file(fname)
        df, techrep_cols = self._prepare_df(df)
        
        try:
            cleaned_df = pd.DataFrame()
            cleaned_df['Aircraft_Register__c'] = df['a/c'].str.strip()
            cleaned_df['Start_Date__c'] = self.normalize_datetime(
                df, 'start date', 'start time(utc)')
            cleaned_df['Release_Date__c'] = self.normalize_datetime(
                df, 'finish date', 'finish time(utc)')
    
            df.fillna('', inplace=True)
    
            cleaned_df['OOS_Total_Time__c'] = self.normalize_oos_time(df, cleaned_df, 'aog time')
    
            cleaned_df['Station__c'] = df['station'].str.strip()
            cleaned_df['Event_Description__c'] = df['defect']
            cleaned_df['Action_Description__c'] = df['rectification action']
    
            for col in df.columns:
                if techrep_cols.get(col, None) is not None:
                    cleaned_df[techrep_cols[col]] = df[col]
            cleaned_df['TechRep_Comments__c'] = df['comments']
            cleaned_df['Reference_Date__c'] = self.get_reference_date(
                fname, len(cleaned_df))
            
            cleaned_df = self._commom_clean(cleaned_df)
        except KeyError as e:
            error = 'FAIL - KEYERROR - NOT RECOGNIZE THE USUAL COLUMN ' + str(e)
            return error, cleaned_df
        
        return "OK", cleaned_df

'''       
class AstanaParser(Parser):

    def __init__(self, root, file_pattern):
        self.root = root
        self.astana_path = '/4 - EMEA/AIR ASTANA'
        super(AstanaParser, self).__init__(
            self.root + self.astana_path, file_pattern)

    @staticmethod
    def _correlate_columns(df):
        c_correlated = {}
        for col in df.columns:
            if col == 'other':
                c_correlated[col] = 'Others__c'
            elif col == 'parts unavailability':
                c_correlated[col] = 'Parts_Unavailability__c'
            elif 'receive embraer disposition' in col:
                c_correlated[col] = 'Time_to_Receive_Embraer_Disposition__c'
            elif 'customer operations' in col:
                c_correlated[col] = 'Customer_Operation__c'
            elif 'time for troubleshooting' in col:
                c_correlated[col] = 'Expected_Time_for_Troubleshooting__c'

        return c_correlated

    def _prepare_df(self, df, fname):
        #df['start date'] = pd.to_datetime(df['start date'], format='%d/%m/%Y')
        #df['start date'] = df['start date'].dt.strftime('%d/%m/%Y')
        df['start date aux'] = df['start date']
        df['start date aux'] = pd.to_datetime(df['start date aux'], format='%d/%m/%Y')
        #df['start date aux'] = df['start date aux'].map(str)
        #df['start date'] = df['start date'].map(str)
        
        df['start date aux'] = df['start date aux'].map(str)
        df['start date aux'] = df['start date aux'].replace('-', '.')
        df['start date aux'] = df['start date aux'].replace(r'([0-9]{1,2}\:[0-9]{1,2}(?:\:[0-9]{1,2})?)', '', regex=True)
        
        #df['start date aux'] = pd.to_datetime(df['start date aux'], format='%d/%m/%Y')
        df['start date aux'] = df['start date aux'].dt.strftime('%d/%m/%Y')
        df['start date aux'] = df['start date aux'].map(str)
        categories = df[['start date', 'start date aux', 'category', 'contrib (%)', 'comments']].copy()
        categories['a/c'] = df['a/c'].ffill().values
        categories['contrib (%)'].map(float).dropna(inplace=True)
        categories['start date'] = categories['start date'].ffill().values
        
        categories['start date aux'] = categories['start date aux'].map(str)
        categories['start date aux'] = categories['start date aux'].replace('-', '.')
        categories['start date aux'] = categories['start date aux'].replace(r'([0-9]{1,2}\:[0-9]{1,2}(?:\:[0-9]{1,2})?)', '', regex=True)
        #categories['start date'] = categories['start date'].replace('-', '.')
        
        categories_pivot = categories.pivot_table(index = ['a/c', 'start date aux'],columns='category', values='contrib (%)')
        ngroups = categories.groupby('category').ngroups
        ac, start, comments = np.asarray(np.split(df[['a/c', 'start date aux', 'comments']].fillna('').values, ngroups)).T
        ac = [' '.join(cell for cell in row) for row in ac]
        comments = [' '.join(cell for cell in row) for row in comments]
        start = [' '.join(cell for cell in row) for row in start]
        start = str(start)
        start = start.split(' ') 
        start = start[0]
        print(start)     
        df_group = pd.DataFrame({'start date aux' : start, 'comments': comments, 'a/c': ac})
        df_group['a/c'] = df_group['a/c'].strip() 
        df_group.to_excel("df_group.xlsx", sheet_name='df_group')
        categories_pivot = pd.merge(categories_pivot, 
                                    df_group,
                                    on=['start date aux', 'a/c'])
        categories_pivot.to_excel("categories_pivot.xlsx", sheet_name='categories_pivot')
        df = df.replace(r'^\s*$', np.nan, regex=True)
        df = df.dropna(axis='rows', subset=[
            'a/c']).drop(columns=['category', 'contrib (%)', 'comments'])
        
        df = pd.merge(df, categories_pivot, on=['a/c', 'start date aux'])
        df.columns = [str(col).strip().lower() for col in df.columns]

        correlated_columns = self._correlate_columns(df)
        for techrep_col in correlated_columns.keys():
            df[techrep_col] = df[techrep_col].apply(
                lambda x: x*100 if x <= 1 else x).fillna('').astype(str)
        return df, correlated_columns


    def get_cleaned_df(self, fname):
        warnings.filterwarnings('ignore')
        print('Iniciando o processamento de dados da AIR ASTANA')
        warnings.filterwarnings('ignore')
        converters = {'Rectification Action': (lambda x: x.upper())}
        df = self.load_file_astana(fname, converters=converters)
        df, techrep_cols = self._prepare_df(df, fname)
        df.to_excel("DF.xlsx", sheet_name='DF')
        
        try:
            cleaned_df = pd.DataFrame()
            cleaned_df['Aircraft_Register__c'] = df['a/c'].str.strip()
            print(df['start date'])
            cleaned_df['Start_Date__c'] = self.normalize_datetime(
                df, 'start date', 'start time(utc)')
            cleaned_df['Release_Date__c'] = self.normalize_datetime(
                df, 'finish date', 'finish time(utc)')
    
            df.fillna('', inplace=True)
    
            cleaned_df['OOS_Total_Time__c'] = self.normalize_oos_time(df, cleaned_df, 'aog time')
    
            cleaned_df['Station__c'] = df['station'].str.strip()
            cleaned_df['Event_Description__c'] = df['defect']
            cleaned_df['Action_Description__c'] = df['rectification action']
    
            for col in df.columns:
                if techrep_cols.get(col, None) is not None:
                    cleaned_df[techrep_cols[col]] = df[col]
            cleaned_df['TechRep_Comments__c'] = df['comments']
            cleaned_df['Reference_Date__c'] = self.get_reference_date(
                fname, len(cleaned_df))
            
            cleaned_df = self._commom_clean(cleaned_df)
        except KeyError as e:
            error = 'FAIL - KEYERROR - NOT RECOGNIZE THE USUAL COLUMN ' + str(e)
            return error, cleaned_df
        
        return "OK", cleaned_df
''' 

class KLMParser(Parser):
    #Get the KLM path
    def __init__(self, root, file_pattern):
        self.root = root
        self.klm_path = '/4 - EMEA/KLM CITYHOPPER'
        super(KLMParser, self).__init__(
            self.root + self.klm_path, file_pattern)


        
    #Get all spreadsheet records
    def get_cleaned_df(self, fname):
        print('Iniciando a importação de dados da KLM CITYHOOPER...')
        warnings.filterwarnings('ignore')
        
        #self.parser_sheet_pirep_klm(fname)
        '''
        conversion of the records of A/c to full format and
        conversion of the records of Remarks to uppercase
        '''
        print('Preparando os dados...')
        converters = {
            'A/C': (lambda x: x if x.startswith('PH') else 'PH-' + x.strip()),
            'Remarks': (lambda x: x.upper())}
        #Read the excel file operator
        df = self.load_file_klm(fname,converters=converters)
        cleaned_df = pd.DataFrame()
        df_filter_pirep_formated = pd.DataFrame(df)
        #Data filtering for E2 Family
        df_filter_pirep_formated = df[df['ac type'] == 'E195E2']
        #Normalizes the field ATA to String
        df_filter_pirep_formated['ata'] = df_filter_pirep_formated['ata'].map(str)
        #Retiring the character 0 from the beginning of the ATA field (RCP format)
        df_filter_pirep_formated['ata'] = df_filter_pirep_formated['ata'].str.lstrip('0')
        df_filter_pirep_formated['ata'] = df_filter_pirep_formated['ata'].map(int)
        ac_format = df_filter_pirep_formated['a/c'].str.split("-", n=1, expand=True)
        df_filter_pirep_formated['a/c_l'] = ac_format[0]
        df_filter_pirep_formated['a/c_r'] = ac_format[1]
        df_filter_pirep_formated['a/c_aux'] = df_filter_pirep_formated['a/c_l'] + df_filter_pirep_formated['a/c_r']
        df_filter_pirep_formated['a/c_aux']= df_filter_pirep_formated['a/c_aux'].str.strip()
        df_filter_pirep_formated['remarks'] = df_filter_pirep_formated['remarks'].str.strip()
        #Normalizes the date to the RCP format
        df_filter_pirep_formated['report date'] = pd.to_datetime(df_filter_pirep_formated['start date'], format='%d-%m-%Y')
        df_filter_pirep_formated['report date'] = df_filter_pirep_formated['report date'].map(str)
        df_filter_pirep_formated['report date'] = df_filter_pirep_formated['report date'].str.strip()
        df_filter = pd.DataFrame(df)
        #Data filtering for E2 Family
        df_filter = df[df['ac type'] == 'E195E2']
        #Normalizes the field ATA to String
        df_filter['ata'] = df_filter['ata'].map(str)
        #Retiring the character 0 from the beginning of the ATA field (RCP format)
        df_filter['ata'] = df_filter['ata'].str.lstrip('0')
        #Normalizes the date to the RCP format
        df_filter['report date'] = pd.to_datetime(df_filter['start date'], format='%d-%m-%Y')
        df_filter['report date'] = df_filter['report date'].map(str)
        df_filter['report date'] = df_filter['report date'].str.strip()
        print('Dados preparados com sucesso!')
        
        print('Lendo as informações de INTER e PIREP')
        #READING THE PIREP AND INTER DATA
        df_filter_pirep = pd.read_excel("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/KLM/INTER_PIREP_KLM.xlsx", sheet_name='PIREP')
        df_filter_inter = pd.read_excel("//flmfs05/vss/suporte_tecnico/FPR/6 - DADOS OPERACIONAIS/1 - OPERADORES/2 - Dados trabalhados/11 - OOS/OOS_SALESFORCE_APP_v2/INTER_PIREP/KLM/INTER_PIREP_KLM.xlsx", sheet_name='INTER')
        print('Dados lidos com sucesso!')
        print('Fazendo a concatenação com PIREP')
        #VLOOKUP to PIREP records (Merge Funcition)
        df_filter = pd.merge(df_filter_pirep, df_filter_pirep_formated, left_on=['DATE_OOS',
                                                                     'ATA_OOS',
                                                                    'REGISTER_OOS',
                                                                    'EVENT_OOS'], 
                                                                right_on=['report date',
                                                                          'ata',
                                                                         'a/c_aux', 
                                                                         'remarks'], 
                                                                          how = 'right')
        #Returns the field a/c to the normal format
        #df_filter['a/c'] = df_filter_pirep_formated['a/c_l'] + '-' + df_filter_pirep_formated['a/c_r']
        print('PIREP concatenado com sucesso!')
        #VLOOKUP to INTERRUPTIONS records (Merge Funcition)
        print('Fazendo a concatenação com INTER')        
        df_filter = pd.merge(df_filter_inter, df_filter, left_on=['DATE_OOS', 
                                                                    'REGISTER_OOS', 
                                                                    'ATA_OOS',
                                                                    'EVENT_OOS'], 
                                                                right_on=['report date', 'a/c', 
                                                                         'ata',
                                                                         'remarks'], 
                                                                          how = 'right')
        print('PIREP concatenado com sucesso!')
        df_filter.fillna('', inplace=True)
        '''
        Concats the field REMARKS with the field CAUSE TEXT (PIREP CAUSE) to create the Event
        Description Field (Full format)
        '''
        df_filter['EVENT_OOS_y'] = '('+'**TYPE**: '+df_filter['type'].map(str)+')' +'\n'+'\n'+ df_filter['EVENT_OOS_y'].map(str) 
        df_filter['EVENT_OOS_y'] = df_filter['EVENT_OOS_y'] + df_filter['remarks'].map(str)+'\n'+'\n'+' **PIREP CAUSE**: '+df_filter['EVENT_PIREP'].map(str)
        df_filter = df_filter.filter(items= ['Problem', 'ACTION_PIREP','a/c', 'start date', 'start time', 
                                             'end date', 'end time', 'out station', 
                                             'ata', 'delta', 'EVENT_OOS_y'])
        
        #Transferring records to the Salesforce dataframe
        try:
            
            cleaned_df['Aircraft_Register__c'] = df_filter['a/c']
            cleaned_df['Start_Date__c'] = self.normalize_datetime(df_filter, 
                                                                  'start date', 'start time')
            cleaned_df['Release_Date__c'] = self.normalize_datetime(df_filter, 
                                                                  'end date', 'end time')
            cleaned_df['OOS_Total_Time__c'] = self.normalize_oos_time(df_filter, 
                                                                  cleaned_df, 'delta')
            cleaned_df['Operator_ATA_Chapter__c'] = df_filter['ata']
            cleaned_df['Event_Description__c'] = df_filter['EVENT_OOS_y']
            cleaned_df['Action_Description__c'] = df_filter['ACTION_PIREP']
            cleaned_df['Inter_ID__c'] = df_filter['Problem']
            cleaned_df['Station__c'] = df_filter['out station'].str.strip()

            cleaned_df['Reference_Date__c'] = self.get_reference_date(fname, 
                                                                      len(cleaned_df))

            cleaned_df = self._commom_clean(cleaned_df)
            cleaned_df.to_excel("CLEANED_DF.xlsx", sheet_name='DF', index=False)
        #If it goes wrong
        except KeyError as e:
            error = 'FAIL - KEYERROR - NOT RECOGNIZE THE USUAL COLUMN ' + str(e)
            return error, cleaned_df
        
        return "OK", cleaned_df
class BelaviaParser(Parser):
    #get the BELAVIA path
    def __init__(self, root, file_pattern):
        self.root =  root
        self.belavia_path = '/4 - EMEA/BELAVIA'
        super(BelaviaParser, self).__init__(
            self.root + self.belavia_path, file_pattern)

    #get all spreadsheet records
    def get_cleaned_df(self, fname):
        
        warnings.filterwarnings('ignore')
        #conversion of the records of AIRCRAFT REGISTRATION NUMBER to full format
        converters = {
            'AIRCRAFT REGISTRATION NUMBER': (lambda x: x if x.startswith('EW') else 'EW-' + x.strip())} #Function LAMBDA to convert the AIRCRAFT REGISTRATION NUMBER
        ac = ['EW-555PO', 'EW-560PO', 'EW-563PO'] #List of E2 Aircrafts
        #read the excel
        df = self.load_file_belavia(fname, converters=converters)
        cleaned_df = pd.DataFrame()
        #data filtering for Unschedule Events and E2 Family
        df_filter = pd.DataFrame(df)
        df_filter = df[df['out of service type'] == 'U']
        df_filter['aircraft_registration_number'] = df_filter['aircraft registration number']
        df_filter_family = pd.DataFrame(df_filter)
        df_filter_family = df_filter_family.query('aircraft_registration_number == @ac')  #query to the filter E2 Family

        #transferring records to the Salesforce dataframe 
        try:
            
            cleaned_df['Aircraft_Register__c'] = df_filter_family['aircraft registration number'].str.strip()
            cleaned_df['Start_Date__c'] = self.normalize_datetime(df_filter_family, 'aircraft removed from service date',
                                                                         'out of service time') #convertion of start date and start time (line 95)
            cleaned_df['Release_Date__c'] = self.normalize_datetime(df_filter_family, 'maintenance released date',
                                                                           'maintenance released time') #convertion of end date and end time (line 95) 

            cleaned_df['OOS_Total_Time__c'] = self.normalize_oos_time(df_filter_family, cleaned_df, None) #convertion of OOS Time to the format decimal (line 147)
            cleaned_df['Operator_ATA_Chapter__c'] = df_filter_family['airline ata system code (discrepancy/symptom)']
            cleaned_df['Event_Description__c'] = df_filter_family['oos event text']
            cleaned_df['Action_Description__c'] = df_filter_family['oos event corrective action text']
            cleaned_df['Station__c'] = df_filter_family['aircraft out of service station'].str.strip()

            cleaned_df['Reference_Date__c'] = self.get_reference_date(fname, len(cleaned_df))   #get referece date (line 197)

            cleaned_df = self._commom_clean(cleaned_df) #dataframe format (line 74)
            
        #if it goes wrong 
        except KeyError as e:
            error = 'FAIL - KEYERROR - NOT RECOFNIZE THE USUAL COLUMN ' + str(e)
            return error, cleaned_df
        
        return "OK", cleaned_df
